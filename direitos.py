import re
from io import BytesIO
from datetime import datetime, timedelta

import pandas as pd
import pdfplumber
import streamlit as st
from openpyxl import load_workbook


st.set_page_config(page_title="Cálculo PMMG", layout="wide")


# =========================================================
# EXTRAÇÃO DO PDF
# =========================================================

def extrair_texto_pdf(pdf_file) -> str:
    texto = ""
    with pdfplumber.open(pdf_file) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text() or ""
            texto += page_text + "\n"
    return texto


def extrair_data_referencia(texto: str):
    padroes = [
        r"Data de referência[:\s]+(\d{2}/\d{2}/\d{4})",
        r"DATA DE REFER[ÊE]NCIA[:\s]+(\d{2}/\d{2}/\d{4})",
        r"AT[ÉE]\s+A\s+DATA\s+DE\s+(\d{2}/\d{2}/\d{4})",
        r"RESUMO DO TEMPO DE SERVIÇO DO MILITAR AT[ÉE] A DATA DE\s+(\d{2}/\d{2}/\d{4})",
    ]

    for padrao in padroes:
        m = re.search(padrao, texto, re.IGNORECASE | re.DOTALL)
        if m:
            try:
                return datetime.strptime(m.group(1), "%d/%m/%Y")
            except ValueError:
                pass
    return None


def extrair_anos_dias_label(texto: str, label_regex: str) -> tuple[int, int]:
    padroes = [
        rf"{label_regex}\s*:\s*(\d+)\s+(\d+)",
        rf"{label_regex}\s+(\d+)\s+(\d+)",
        rf"{label_regex}.*?(\d+)\s+anos.*?(\d+)\s+dias",
    ]

    for padrao in padroes:
        m = re.search(padrao, texto, re.IGNORECASE | re.DOTALL)
        if m:
            return int(m.group(1)), int(m.group(2))

    return 0, 0


def anos_dias_para_dias(anos: int, dias: int) -> int:
    return anos * 365 + dias


def extrair_nome(texto: str) -> str:
    m = re.search(r"NOME:\s*(.+)", texto)
    return m.group(1).strip() if m else ""


def extrair_posto(texto: str) -> str:
    m = re.search(r"POSTO OU GRADUAÇÃO:\s*(.+?)(?:\s+NÚMERO PM:|$)", texto)
    return m.group(1).strip() if m else ""


# =========================================================
# REGRA DE NEGÓCIO
# =========================================================

def calcular_data_marco(data_ref: datetime, total_dias_ajustado: int, marco_dias: int) -> datetime:
    excesso = total_dias_ajustado - marco_dias
    return data_ref - timedelta(days=excesso)


def calcular_data_quinquenio(data_ref: datetime, total_dias_ajustado: int, numero: int) -> datetime:
    marco = numero * 1825
    return calcular_data_marco(data_ref, total_dias_ajustado, marco)


def calcular_data_trintenario(data_ref: datetime, total_dias_ajustado: int) -> datetime:
    return calcular_data_marco(data_ref, total_dias_ajustado, 10950)


def extrair_ferias_anuais_nao_gozadas_simples(texto: str) -> int:
    """
    Soma os dias simples da seção:
    FÉRIAS ANUAIS – NÃO GOZADAS
    """
    linhas = texto.splitlines()
    captura = False
    total = 0

    for linha in linhas:
        linha_up = linha.upper().strip()

        if "FÉRIAS ANUAIS" in linha_up and "NÃO GOZADAS" in linha_up:
            captura = True
            continue

        if captura and (
            "RESUMO DO TEMPO DE SERVIÇO" in linha_up
            or "FÉRIAS ANUAIS CONTADAS DE FORMA SIMPLES" in linha_up
            or "FÉRIAS ANUAIS CONTADAS EM DOBRO" in linha_up
        ):
            break

        if captura:
            m = re.search(r"^\s*(\d{4})\s+(\d+)\s+(DOBRO|SIMPLES)?\s*$", linha, re.IGNORECASE)
            if m:
                dias = int(m.group(2))
                total += dias

    return total


def diff_ymd(data_inicial: datetime, data_final: datetime) -> tuple[int, int, int]:
    dias = (data_final.date() - data_inicial.date()).days
    anos = dias // 365
    resto = dias % 365
    meses = resto // 30
    dias_restantes = resto % 30
    return anos, meses, dias_restantes


# =========================================================
# PREENCHIMENTO DO MODELO EXCEL
# =========================================================

def preencher_modelo_excel(
    caminho_modelo: str,
    nome: str,
    posto: str,
    data_ref: datetime,
    total_sirh_dias: int,
    ferias_nao_gozadas_simples: int,
    total_ajustado_dias: int,
    data_6qq: datetime,
    data_trint: datetime,
) -> BytesIO:
    wb = load_workbook(caminho_modelo)
    ws = wb.active

    ws["B2"] = nome
    ws["B3"] = posto
    ws["B4"] = data_ref.strftime("%d/%m/%Y")
    ws["B6"] = total_sirh_dias
    ws["B7"] = ferias_nao_gozadas_simples
    ws["B8"] = total_ajustado_dias
    ws["B10"] = data_6qq.strftime("%d/%m/%Y")
    ws["B11"] = data_trint.strftime("%d/%m/%Y")

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# =========================================================
# INTERFACE
# =========================================================

st.title("Cálculo de Quinquênio / Trintenário - PMMG")
st.write("Base: PDF do SIRH, com exclusão apenas das férias anuais não gozadas.")

pdf_file = st.file_uploader("Envie o PDF do SIRH", type=["pdf"])

modelo_excel = "modelo_planilha_oficial.xlsx"

if pdf_file is not None:
    texto = extrair_texto_pdf(pdf_file)

    texto_upper = texto.upper()
    marcadores = [
        "RESUMO DO TEMPO DE SERVIÇO",
        "TOTAL DE ANOS DE SERVIÇO",
        "TOTAL DE ACRÉSCIMOS LEGAIS",
    ]
    if not any(m in texto_upper for m in marcadores):
        st.error("O arquivo enviado não parece ser um PDF de contagem de tempo do SIRH.")
        st.stop()

    nome = extrair_nome(texto)
    posto = extrair_posto(texto)
    data_ref = extrair_data_referencia(texto)

    total_anos, total_dias_rest = extrair_anos_dias_label(
        texto,
        r"TOTAL DE ANOS DE SERVIÇO"
    )

    if data_ref is None:
        st.error("Não foi possível localizar a data de referência no PDF.")
        st.stop()

    if total_anos == 0 and total_dias_rest == 0:
        st.error("Não foi possível localizar o campo 'TOTAL DE ANOS DE SERVIÇO' no PDF.")
        st.stop()

    total_sirh_dias = anos_dias_para_dias(total_anos, total_dias_rest)

    ferias_nao_gozadas_simples = extrair_ferias_anuais_nao_gozadas_simples(texto)
    ferias_nao_gozadas_em_dobro = ferias_nao_gozadas_simples * 2

    total_ajustado_dias = total_sirh_dias - ferias_nao_gozadas_em_dobro

    data_6qq = calcular_data_quinquenio(data_ref, total_ajustado_dias, 6)
    data_trint = calcular_data_trintenario(data_ref, total_ajustado_dias)

    st.subheader("Auditoria")
    auditoria = pd.DataFrame(
        [
            ["Nome", nome],
            ["Posto / Graduação", posto],
            ["Data de referência", data_ref.strftime("%d/%m/%Y")],
            ["Total do SIRH (anos)", total_anos],
            ["Total do SIRH (dias excedentes)", total_dias_rest],
            ["Total do SIRH em dias", total_sirh_dias],
            ["Férias anuais não gozadas (simples)", ferias_nao_gozadas_simples],
            ["Férias anuais não gozadas (em dobro)", ferias_nao_gozadas_em_dobro],
            ["Total ajustado em dias", total_ajustado_dias],
        ],
        columns=["Campo", "Valor"]
    )
    st.dataframe(auditoria, use_container_width=True)

    st.subheader("Resultado principal")
    st.write(f"**6º Quinquênio:** {data_6qq.strftime('%d/%m/%Y')}")
    st.write(f"**Adicional trintenário:** {data_trint.strftime('%d/%m/%Y')}")

    st.subheader("Demais quinquênios")
    resultados = []
    for i in range(1, 10):
        dt = calcular_data_quinquenio(data_ref, total_ajustado_dias, i)
        status = "Adquirido" if dt.date() <= data_ref.date() else "Futuro"
        anos, meses, dias = diff_ymd(data_ref, dt) if dt > data_ref else (0, 0, 0)

        resultados.append(
            {
                "Quinquênio": f"{i}º",
                "Data": dt.strftime("%d/%m/%Y"),
                "Status": status,
                "Faltam": f"{anos}a {meses}m {dias}d" if status == "Futuro" else "-"
            }
        )

    st.dataframe(pd.DataFrame(resultados), use_container_width=True)

    try:
        excel_bytes = preencher_modelo_excel(
            caminho_modelo=modelo_excel,
            nome=nome or "Sem nome",
            posto=posto or "Não informado",
            data_ref=data_ref,
            total_sirh_dias=total_sirh_dias,
            ferias_nao_gozadas_simples=ferias_nao_gozadas_simples,
            total_ajustado_dias=total_ajustado_dias,
            data_6qq=data_6qq,
            data_trint=data_trint,
        )

        st.download_button(
            label="Baixar planilha preenchida",
            data=excel_bytes,
            file_name="planilha_preenchida.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except FileNotFoundError:
        st.warning("Modelo Excel não encontrado na raiz do projeto: modelo_planilha_oficial.xlsx")
